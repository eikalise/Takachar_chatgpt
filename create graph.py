import os
import random
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook

files = [f for f in os.listdir('/Users/EIK/Downloads/Takachar-GPT-main/Final_Excels_Alise') if f.endswith('.xlsx')]

selected_files = random.sample(files, 6)

fig, axs = plt.subplots(2, 3, figsize=(20, 12))

for i, file in enumerate(selected_files):
    # construct the file path
    file_path = os.path.join('/Users/EIK/Downloads/Takachar-GPT-main/Final_Excels_Alise', file)

    # load the workbook and get the first sheet name
    wb = load_workbook(filename=file_path)
    sheet_name = wb.sheetnames[0]

    # read the data into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # create the plot
    row = i // 3
    col = i % 3
    sns.scatterplot(data=df, x="delta_times", y="combustor_temp", hue="State", palette="deep", ax=axs[row, col])

    # add title using file name
    axs[row, col].set_title(f"delta_times vs combustor_temp for {file}")

plt.tight_layout()
plt.show()